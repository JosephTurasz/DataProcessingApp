from __future__ import annotations

import re

from openpyxl.styles import PatternFill

# ---- Brief detection --------------------------------------------------------

BRIEF_TYPES = {
    "client/sub-client name": "TDG Brief v2",
    "client to be billed":    "TDG Brief v1",
}

HEADER_RENAME_RULES = [
    ("Machineable",        "Machinability"),
    ("Weight of Pack",     "Item Weight"),
    ("Number of Items",    "No. Items"),
    ("Date of Collection", "Collection Date"),
]

JOB_HEADER_VARIANTS    = {"Job Reference", "Job Name/Reference"}
CLIENT_HEADER_VARIANTS = {"client/sub-client name", "client to be billed"}

# ---- BAG file mapping -------------------------------------------------------

# Maps output header display names → ImportOption keys in a .BAG.txt file
BAG_FIELD_MAP = {
    "Client":          "Client",
    "Job Name":        "JobRef",
    "Collection Date": "CollectionDate",
    "Accurate Weight": "ItemWeight",
    "Item Weight":     "ItemWeight",
    "Container":       "ContainerType",
}

# ---- Regex patterns ---------------------------------------------------------

SPACE_RUN_RE        = re.compile(r" {5,}")
BRACKET_RE          = re.compile(r"\s*[\(\[][^\)\]]*[\)\]]")
NUMERIC_RE          = re.compile(r"\d+(?:\.\d+)?")
TRAY_FILL_RE        = re.compile(r"\btray\s+fill\b", re.IGNORECASE)
DIGITS_RE           = re.compile(r"\d+")
POSTCOM_RE          = re.compile(r"POSTCOM\s*-\s*Collection\s+Address\s+(\w+\s+\w+)", re.IGNORECASE)
MAX_IN_CONTAINER_RE = re.compile(r"\bmax\s+(\d+)", re.IGNORECASE)
WORD_RE             = re.compile(r"\w+")

# ---- Date parsing -----------------------------------------------------------

DATE_PARSE_FMTS = [
    "%d/%m/%Y", "%d/%m/%y",
    "%Y-%m-%d", "%d-%m-%Y",
    "%Y/%m/%d",
    "%d.%m.%Y", "%d.%m.%y",
]

# ---- Excel cell fills -------------------------------------------------------

GREEN_FILL  = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
RED_FILL    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FF8000", end_color="FF8000", fill_type="solid")

# ---- Column classification --------------------------------------------------

NUMERIC_COLS = frozenset({
    "Max Container Fill", "Container Fill", "Item Weight",
    "Accurate Weight", "Maximum Container Weight", "No. Items",
})

GROUP_COL_NAMES = ("Service", "Format", "Additional Service", "Machinability")

# ---- Service-group validation sections --------------------------------------
#
# Each entry: (section_name, {brief_kw: data_kw}, must_match, symmetric)
#   must_match — True: brief must contain a keyword or the section fails.
#   symmetric  — True: no keyword in brief is only valid if data also has none.
# Keywords are sorted longest-first at use-time so specific phrases beat short ones.

SERVICE_SECTIONS: list[tuple[str, dict[str, str], bool, bool]] = [
    ("Format", {
        "large letter": "large letter",
        "lrg ltr":      "large letter",
        "letter":       "letter",
        "ltr":          "letter",
        "ll":           "large letter",
    }, True, False),
    ("Machinability", {
        "mailmark": "mailmark",
        "mm":       "mailmark",
        "manual":   "manual",
        "man":      "manual",
    }, True, False),
    ("Mail Category", {
        "advertising catalogue": "catalogue",
        "advertising mail":      "advertising",
        "catalogue mail":        "catalogue",
        "partially addressed":   "partially addressed",
        "admail":                "advertising",
        "adv mail":              "advertising",
        "cat":                   "catalogue",
        "pa":                    "partially addressed",
    }, True, False),
    ("Service", {
        "economy": "economy",
        "eco":     "economy",
    }, False, True),
]
