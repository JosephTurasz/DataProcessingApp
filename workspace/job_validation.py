from __future__ import annotations

import os
import re
from datetime import date, datetime
from difflib import SequenceMatcher
from functools import lru_cache

import openpyxl
from openpyxl.styles import Alignment, Font
from PySide6.QtWidgets import (
    QDialog, QLabel, QPushButton, QScrollArea, QVBoxLayout, QWidget,
)

from config.constants.job_validation import (
    BAG_FIELD_MAP, BRACKET_RE, BRIEF_TYPES, CLIENT_HEADER_VARIANTS,
    DATE_PARSE_FMTS, DIGITS_RE, GREEN_FILL, GROUP_COL_NAMES,
    HEADER_RENAME_RULES, JOB_HEADER_VARIANTS, MAX_IN_CONTAINER_RE,
    NUMERIC_COLS, NUMERIC_RE, ORANGE_FILL, POSTCOM_RE, RED_FILL,
    SERVICE_SECTIONS, SPACE_RUN_RE, TRAY_FILL_RE, WORD_RE,
)
from processing.loading import load_bag, load_cpr, load_labels
from workspace.base import BaseWorkflow


def _normalize_cell(val):
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    if not isinstance(val, str):
        return val
    m = SPACE_RUN_RE.search(val)
    return val[:m.start()] if m else val


def _clean_header(val):
    if not isinstance(val, str):
        return val
    return BRACKET_RE.sub("", val).strip()


def _normalise_date(val) -> str | None:
    """Return val as DD/MM/YYYY string, trying common formats. Returns None if blank."""
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    s = str(val or "").strip()
    if not s:
        return None
    for fmt in DATE_PARSE_FMTS:
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return s


def _word_fuzzy_match(w1: str, w2: str) -> bool:
    """True if two words are considered equivalent.

    Handles plurals/truncations via prefix ('tray'/'trays').
    """
    if w1 == w2:
        return True
    short, long_ = (w1, w2) if len(w1) <= len(w2) else (w2, w1)
    return long_.startswith(short)


def _words_match(a: str, b: str) -> bool:
    """True if every word in the shorter string fuzzy-matches a word in the longer.

    Handles truncated names ('Cashmere - T' / 'Cashmere Centre - T'),
    plurals ('Tray' / 'Trays'), and abbreviations ('Bk' / 'Book').
    """
    wa = WORD_RE.findall(a.lower())
    wb = WORD_RE.findall(b.lower())
    if not wa or not wb:
        return False
    shorter, longer = (wa, wb) if len(wa) <= len(wb) else (wb, wa)
    return all(any(_word_fuzzy_match(sw, lw) for lw in longer) for sw in shorter)


# ---- Service-group validation -----------------------------------------------

@lru_cache(maxsize=256)
def _kw_pattern(keyword: str) -> re.Pattern:
    return re.compile(r"\b" + re.escape(keyword) + r"\b", re.IGNORECASE)


def _kw_in(text: str, keyword: str) -> bool:
    return bool(_kw_pattern(keyword).search(text))


# Pre-compute per-section: sorted (kw, data_kw) pairs (longest first) + unique data kws
_COMPILED_SECTIONS: list[tuple[str, list[tuple[str, str]], bool, bool, frozenset[str]]] = [
    (name, sorted(kw_map.items(), key=lambda x: len(x[0]), reverse=True),
     must_match, symmetric, frozenset(kw_map.values()))
    for name, kw_map, must_match, symmetric in SERVICE_SECTIONS
]


def _validate_service_group(brief_text: str, data_text: str) -> list[str]:
    """Return a list of failed section names (empty = all passed)."""
    failed = []
    for section_name, sorted_kws, must_match, symmetric, data_kws in _COMPILED_SECTIONS:
        matched_data_kw = None
        for kw, data_kw in sorted_kws:
            if _kw_in(brief_text, kw):
                matched_data_kw = data_kw
                break

        if matched_data_kw is None:
            if must_match:
                failed.append(section_name)
            elif symmetric and any(_kw_in(data_text, dk) for dk in data_kws):
                failed.append(section_name)
            continue

        if not _kw_in(data_text, matched_data_kw):
            failed.append(section_name)

    return failed


# -----------------------------------------------------------------------------

def _apply_po_prefix(all_rows: list, rows_data: list, job_col_idx: int) -> None:
    """If a 'PO TO ADD TO DOCKETHUB' cell exists, prepend its value to each job name."""
    po_value = _find_right_of(all_rows, "po to add to dockethub")
    if not po_value or not str(po_value).strip():
        return
    for row_data in rows_data:
        current = str(row_data[job_col_idx] or "").strip()
        row_data[job_col_idx] = f"{po_value} {current}".strip() if current else po_value


def _to_numeric(val):
    """Extract the first number from val. Returns int if whole, float otherwise, or None."""
    if isinstance(val, (int, float)):
        return val
    m = NUMERIC_RE.search(str(val or ""))
    if not m:
        return None
    n = float(m.group())
    return int(n) if n == int(n) else n


def _extract_container_fill(val):
    """Return the first integer from the 'Tray Fill ...' line, or val unchanged."""
    if not isinstance(val, str):
        return val
    for line in val.splitlines():
        if TRAY_FILL_RE.search(line):
            m = DIGITS_RE.search(line)
            if m:
                return int(m.group())
    return val


def _find_right_of(all_rows, label: str, *, partial: bool = False):
    """Return the first non-blank cell value to the right of the cell matching label."""
    label_lower = label.lower()
    for row in all_rows:
        for ci, cell_val in enumerate(row):
            s = str(cell_val or "").strip().lower()
            matched = (label_lower in s) if partial else (s == label_lower)
            if matched:
                for right_col in range(ci + 1, len(row)):
                    rv = row[right_col]
                    if rv is not None and str(rv).strip():
                        return _normalize_cell(rv)
                return None
    return None


def _has_line_break(val) -> bool:
    return isinstance(val, str) and ("\n" in val or "\r" in val)


def _col_idx(header_values: list, name: str) -> int | None:
    return next((i for i, hv in enumerate(header_values) if str(hv or "") == name), None)


def _detect_brief_type(all_rows: list) -> tuple[str | None, int | None]:
    """Single pass over all_rows; returns (brief_type, header_row_idx).

    header_row_idx is only set for TDG briefs, where the sentinel cell IS the
    header row.  For all other types it is None — the dedicated parser locates
    its own header internally.
    """
    for i, row in enumerate(all_rows):
        col_a = str(row[0] or "").strip()
        if bt := BRIEF_TYPES.get(col_a.lower()):
            return bt, i
        for cell_val in row:
            s = str(cell_val or "").strip()
            if s.lower() == "job name:":
                return "Pureprint Brief", None
            if s.upper() == "JOB REFERENCE":
                return "Scotts Brief", None
    return None, None


def _parse_tdg_headers(row: tuple) -> tuple[list[str], list]:
    """Extract logical names and cleaned display values from a TDG sentinel row."""
    headers: list[str] = []
    header_values: list = []
    for val in row:
        stripped = str(val or "").strip()
        if not stripped:
            break
        headers.append(stripped)
        header_values.append(_clean_header(_normalize_cell(val)))
    return headers, header_values


class _JobSelectionDialog(QDialog):
    def __init__(self, job_names: list[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Job to Validate")
        self.selected_idx: int | None = None
        self.resize(340, 420)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Select a job to validate:"))

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        btn_layout = QVBoxLayout(container)
        btn_layout.setSpacing(6)

        for i, name in enumerate(job_names):
            btn = QPushButton(str(name))
            btn.setMinimumHeight(38)
            btn.clicked.connect(lambda checked, idx=i: self._select(idx))
            btn_layout.addWidget(btn)

        btn_layout.addStretch()
        scroll.setWidget(container)
        layout.addWidget(scroll)

        cancel = QPushButton("Cancel")
        cancel.clicked.connect(self.reject)
        layout.addWidget(cancel)

    def _select(self, idx: int):
        self.selected_idx = idx
        self.accept()


class JobValidation(BaseWorkflow):
    def run(self, checked: bool = False):
        infile = self.mw.ask_open_file(
            "Select a brief",
            "Excel Files (*.xlsx)",
        )
        if not infile:
            return
        self.busy("Job Validation", "Reading brief…",
                  lambda: self._parse_brief(infile),
                  on_done=lambda result: self._on_brief_parsed(infile, result))

    # ------------------------------------------------------------------
    # Background: parse the brief Excel file
    # ------------------------------------------------------------------
    def _parse_brief(self, infile: str) -> tuple:
        wb = openpyxl.load_workbook(infile, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        brief_type, header_list_idx = _detect_brief_type(all_rows)
        if brief_type is None:
            raise ValueError("Could not identify brief type. Is this a TDG, Pureprint, or Scotts brief?")

        if brief_type == "Scotts Brief":
            hv, rd, ji, bt = self._parse_scotts_brief(all_rows)
            _apply_po_prefix(all_rows, rd, ji)
            return hv, rd, ji, bt

        if brief_type == "Pureprint Brief":
            def _r(label, *, partial=False):
                return _find_right_of(all_rows, label, partial=partial)

            additional_parts = [
                _r("Ad Mail/General Mail/Business Mail/Catalogue Mail:"),
                _r("Partially Addressed:"),
                _r("Magazine Subscription:"),
            ]
            additional_service = " / ".join(
                str(p) for p in additional_parts if p is not None and str(p).strip()
            )

            container_fill = _to_numeric(_r("If Trays please specify tray fill", partial=True))
            item_weight = _to_numeric(_r("Item weight (g):"))
            target_weight = (
                container_fill * item_weight
                if container_fill is not None and item_weight is not None
                else None
            )

            pp_headers = [
                "Poster", "Client", "Job Name", "Service", "Format",
                "Additional Service", "Machinability", "Container",
                "Container Fill", "Accurate Weight", "Maximum Container Weight",
                "Split Release", "No. Items", "Collection Date", "JIC Opt-In",
            ]
            pp_row = [
                _r("Mailing House Name:"),
                _r("Client Name:"),
                _r("Job Name:"),
                _r("Service:"),
                _r("Item format:"),
                additional_service or None,
                None,
                _r("Mail presentation:"),
                container_fill,
                item_weight,
                target_weight,
                None,
                _to_numeric(_r("Number of items:")),
                _r("Collection Date:"),
                _r("JIC Opt in or Opt out:"),
            ]
            pp_rows = [pp_row]
            _apply_po_prefix(all_rows, pp_rows, pp_headers.index("Job Name"))
            return pp_headers, pp_rows, pp_headers.index("Job Name"), brief_type

        # TDG brief — parse headers from the sentinel row
        # headers       — stripped logical names used for column matching
        # header_values — cleaned display values written to the output
        assert header_list_idx is not None
        headers, header_values = _parse_tdg_headers(all_rows[header_list_idx])

        # Find and rename the job and client columns
        job_col_idx: int | None = None
        for i, h in enumerate(headers):
            if h in JOB_HEADER_VARIANTS:
                headers[i] = "Job Name"
                header_values[i] = "Job Name"
                job_col_idx = i
            elif h.lower() in CLIENT_HEADER_VARIANTS:
                headers[i] = "Client"
                header_values[i] = "Client"

        if job_col_idx is None:
            variants = " or ".join(f'"{v}"' for v in sorted(JOB_HEADER_VARIANTS))
            raise ValueError(f"Could not find {variants} in the header row.")

        # Collect data rows — stop at first blank in the job column
        rows_data: list[list] = []
        for row in all_rows[header_list_idx + 1:]:
            job_val = str(row[job_col_idx] or "").strip() if len(row) > job_col_idx else ""
            if not job_val:
                break
            row_data = [
                _normalize_cell(row[ci]) if len(row) > ci else None
                for ci in range(len(headers))
            ]
            rows_data.append(row_data)

        # TDG Brief v1: Container Fill is absent — insert it and populate.
        # First check if the Container cell contains "Max <n>" (e.g. "Trays (Max 75)");
        # if not, fall back to the cell below "Additional Information".
        if brief_type == "TDG Brief v1":
            container_idx = next(
                (i for i, hv in enumerate(header_values) if str(hv or "").strip() == "Container"), None
            )
            if container_idx is not None:
                insert_at = container_idx + 1
                headers.insert(insert_at, "Max Container Fill")
                header_values.insert(insert_at, "Max Container Fill")
                for row_data in rows_data:
                    row_data.insert(insert_at, None)
                if job_col_idx >= insert_at:
                    job_col_idx += 1

                for row_data in rows_data:
                    container_val = str(row_data[container_idx] or "")
                    m = MAX_IN_CONTAINER_RE.search(container_val)
                    if m:
                        row_data[insert_at] = int(m.group(1))

                if all(row_data[insert_at] is None for row_data in rows_data):
                    max_fill_value = None
                    found = False
                    for row_idx, row in enumerate(all_rows):
                        for ci, cell_val in enumerate(row):
                            if str(cell_val or "").strip().lower() == "additional information":
                                next_row = all_rows[row_idx + 1] if row_idx + 1 < len(all_rows) else []
                                raw = next_row[ci] if ci < len(next_row) else None
                                max_fill_value = _extract_container_fill(_normalize_cell(raw))
                                found = True
                                break
                        if found:
                            break
                    if max_fill_value is not None:
                        for row_data in rows_data:
                            row_data[insert_at] = max_fill_value

        # TDG Brief v2: Container Fill column is already present in the brief.

        # Insert "Maximum Container Weight" (Container Fill × Pack Weight) — always present
        # so the CPR value can always be written to it.
        pack_w_idx = next(
            (i for i, hv in enumerate(header_values)
             if "Weight of Pack" in str(hv or "") or "Item Weight" in str(hv or "")), None
        )
        fill_idx = next(
            (i for i, hv in enumerate(header_values) if "Container Fill" in str(hv or "")), None
        )
        if pack_w_idx is not None:
            twc_at = pack_w_idx + 1
        elif fill_idx is not None:
            twc_at = fill_idx + 1
        else:
            twc_at = len(headers)
        headers.insert(twc_at, "Maximum Container Weight")
        header_values.insert(twc_at, "Maximum Container Weight")
        if job_col_idx >= twc_at:
            job_col_idx += 1
        if fill_idx is not None and fill_idx >= twc_at:
            fill_idx += 1
        for row_data in rows_data:
            row_data.insert(twc_at, None)
            if pack_w_idx is not None and fill_idx is not None:
                pack_w = _to_numeric(row_data[pack_w_idx])
                fill = _to_numeric(row_data[fill_idx])
                if pack_w is not None and fill is not None:
                    row_data[twc_at] = pack_w * fill

        # Insert "Poster" as the first column — first non-blank cell to the right
        # of any cell containing "company name".
        poster_value = _find_right_of(all_rows, "company name", partial=True)
        if poster_value:
            m = POSTCOM_RE.search(str(poster_value))
            if m:
                poster_value = m.group(1)

        headers.insert(0, "Poster")
        header_values.insert(0, "Poster")
        job_col_idx += 1
        for row_data in rows_data:
            row_data.insert(0, poster_value)

        # Apply final display renames to header_values
        for i, hv in enumerate(header_values):
            s = str(hv or "")
            for substr, new_name in HEADER_RENAME_RULES:
                if substr in s:
                    header_values[i] = new_name
                    break

        _apply_po_prefix(all_rows, rows_data, job_col_idx)
        return header_values, rows_data, job_col_idx, brief_type

    # ------------------------------------------------------------------
    # Scotts Brief parser
    # ------------------------------------------------------------------
    @staticmethod
    def _parse_scotts_brief(all_rows: list) -> tuple:
        # Header row: first row containing "Sortation File Volume".
        header_row_idx = next(
            (i for i, row in enumerate(all_rows)
             if any("Sortation File Volume" in str(v or "") for v in row)),
            None,
        )
        if header_row_idx is None:
            raise ValueError("Could not find header row in Scotts Brief.")

        header_row = all_rows[header_row_idx]

        def _find(partial: str) -> int | None:
            pl = partial.lower()
            return next((ci for ci, v in enumerate(header_row) if pl in str(v or "").lower()), None)

        def _find_exact(label: str) -> int | None:
            return next(
                (ci for ci, v in enumerate(header_row) if str(v or "").strip().lower() == label.lower()),
                None,
            )

        job_ref_col   = _find("job reference")
        mailing_col   = _find("mailing house")
        sortation_col = _find("sortation file volume")
        service_col   = _find_exact("service")
        format_col    = _find_exact("format")
        machine_col   = _find("machine")
        tray_col      = _find("max tray fill")
        weight_col    = _find("item weight")
        collect_col   = _find("collection date")

        if job_ref_col is None:
            raise ValueError("Could not find 'Job Reference' column in Scotts Brief.")

        # Build mh_by_row: maps row index → Poster value.
        # Each non-blank mailing house cell is a group anchor; subsequent rows belong
        # to that group until the next anchor is reached, or until 4 consecutive rows
        # with no mailing house value AND no job reference are seen (end of data).
        mh_by_row: dict[int, str | None] = {}
        if mailing_col is not None:
            anchors: list[tuple[int, str | None]] = []
            for row_i in range(header_row_idx + 1, len(all_rows)):
                row = all_rows[row_i]
                mh_cell = row[mailing_col] if mailing_col < len(row) else None
                if str(mh_cell or "").strip():
                    anchors.append((row_i, _normalize_cell(mh_cell)))

            for anchor_idx, (anchor_row_i, poster) in enumerate(anchors):
                next_anchor = anchors[anchor_idx + 1][0] if anchor_idx + 1 < len(anchors) else len(all_rows)
                consecutive_blank = 0
                for row_i in range(anchor_row_i, next_anchor):
                    row = all_rows[row_i]
                    mh_str  = str((row[mailing_col] if mailing_col < len(row) else None) or "").strip()
                    job_str = str((row[job_ref_col]  if job_ref_col  < len(row) else None) or "").strip()
                    if not mh_str and not job_str:
                        consecutive_blank += 1
                        if consecutive_blank >= 4:
                            break
                    else:
                        consecutive_blank = 0
                        mh_by_row[row_i] = poster

        scotts_headers = [
            "Poster", "Job Name", "Service", "Format",
            "Container Fill", "Accurate Weight", "Maximum Container Weight", "No. Items", "Collection Date",
        ]

        def _cell(row, col):
            return _normalize_cell(row[col]) if col is not None and col < len(row) else None

        rows_data: list[list] = []
        for row_i, row in enumerate(all_rows):
            if row_i not in mh_by_row:
                continue
            job_ref = str(row[job_ref_col] or "").strip() if job_ref_col < len(row) else ""
            if not job_ref:
                continue

            service = _cell(row, service_col)
            machine = _cell(row, machine_col)
            service_out = f"{machine} / {service}" if machine and service else service

            tray_fill = _to_numeric(_cell(row, tray_col))
            item_weight = _to_numeric(_cell(row, weight_col))
            max_container_weight = (
                tray_fill * item_weight
                if tray_fill is not None and item_weight is not None
                else None
            )
            rows_data.append([
                mh_by_row[row_i],
                _cell(row, job_ref_col),
                service_out,
                _cell(row, format_col),
                tray_fill,
                item_weight,
                max_container_weight,
                _to_numeric(_cell(row, sortation_col)),
                _cell(row, collect_col),
            ])

        if not rows_data:
            raise ValueError("No data rows found in Scotts Brief.")

        return scotts_headers, rows_data, scotts_headers.index("Job Name"), "Scotts Brief"

    # ------------------------------------------------------------------
    # UI thread: pick job row, ask for validation files, kick off write
    # ------------------------------------------------------------------
    def _on_brief_parsed(self, infile: str, result: tuple):
        header_values, rows_data, job_col_idx, brief_type = result

        self.info(f"[BRIEF] {brief_type} detected.", "green")

        if not rows_data:
            self.warn("Job Validation", "No job rows found in the brief.")
            return

        if len(rows_data) == 1:
            selected_row = rows_data[0]
        else:
            job_names = [str(row[job_col_idx] or "") for row in rows_data]
            dlg = _JobSelectionDialog(job_names, parent=self.mw)
            if dlg.exec() != QDialog.Accepted or dlg.selected_idx is None:
                return
            selected_row = rows_data[dlg.selected_idx]

        validation_files = self.mw.ask_open_files(
            "Select files to validate",
            "Validation Files (*.BAG.txt *.CPR.txt *.LABELS.txt);;All Files (*)",
        )
        if not validation_files:
            return

        job_name = str(selected_row[job_col_idx] or "")
        out_path = os.path.join(os.path.dirname(infile), f"{job_name} Validation File.xlsx")

        self.busy("Job Validation", "Writing validation file…",
                  lambda: self._write_output(out_path, header_values, selected_row, validation_files),
                  on_done=lambda _: self._on_written(out_path, job_name))

    # ------------------------------------------------------------------
    # Background: build and save the validation Excel workbook
    # ------------------------------------------------------------------
    def _write_output(
        self,
        out_path: str,
        header_values: list,
        selected_row: list,
        validation_files: list[str],
    ):
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active

        out_ws.append(["Source"] + header_values)
        for cell in out_ws[1]:
            cell.font = Font(bold=True)
            if _has_line_break(cell.value):
                cell.alignment = Alignment(wrap_text=True)

        for i, hv in enumerate(header_values):
            if str(hv or "") in NUMERIC_COLS:
                selected_row[i] = _to_numeric(selected_row[i])
            elif str(hv or "") == "Collection Date":
                selected_row[i] = _normalise_date(selected_row[i])

        out_ws.append(["Brief"] + selected_row)
        row_num = out_ws.max_row
        out_ws.cell(row=row_num, column=1).font = Font(bold=True)
        for excel_col, val in enumerate(selected_row, start=2):
            if _has_line_break(val):
                out_ws.cell(row=row_num, column=excel_col).alignment = Alignment(wrap_text=True)

        bag_files    = [f for f in validation_files if f.upper().endswith(".BAG.TXT")]
        cpr_files    = [f for f in validation_files if f.upper().endswith(".CPR.TXT")]
        labels_files = [f for f in validation_files if f.upper().endswith(".LABELS.TXT")]
        val_row = [None] * len(header_values)

        def _set(col_name: str, value) -> None:
            idx = _col_idx(header_values, col_name)
            if idx is not None:
                val_row[idx] = value

        bag_df = None
        if bag_files:
            try:
                bag_opts, bag_df = load_bag(bag_files[0])
                for i, hv in enumerate(header_values):
                    opt_key = BAG_FIELD_MAP.get(str(hv or ""))
                    if opt_key:
                        val_row[i] = bag_opts.get(opt_key)
            except Exception as e:
                self.info(f"[BAG] Could not read BAG file: {e}", "yellow")

        if cpr_files:
            try:
                cpr_data = load_cpr(cpr_files[0])

                total_items = cpr_data.get("Total Items Processed", "")
                if total_items:
                    _set("No. Items", _to_numeric(total_items.replace(",", "")))

                max_bag_weight = cpr_data.get("Maximum Bag Weight", "")
                if max_bag_weight:
                    _set("Maximum Container Weight", _to_numeric(max_bag_weight.replace(",", "")))

                service_name = cpr_data.get("Service Selected", "")
                if service_name:
                    svc = self.mw.s.mailsort_services_repo.get_by_service_name(service_name)
                    if svc:
                        sortation = str(svc.get("sortation", "") or "").strip()
                        machinability = str(svc.get("machinability", "") or "").strip()
                        service_str = " / ".join(filter(None, [sortation, machinability]))
                        svc_fields = {
                            "Service":            service_str,
                            "Format":             str(svc.get("format", "") or "").strip(),
                            "Additional Service": str(svc.get("mail_category", "") or "").strip(),
                        }
                        for i, hv in enumerate(header_values):
                            v = svc_fields.get(str(hv or ""))
                            if v is not None:
                                val_row[i] = v
            except Exception as e:
                self.info(f"[CPR] Could not read CPR file: {e}", "yellow")

        if labels_files:
            try:
                labels_data = load_labels(labels_files[0])
                poster = labels_data.get("Poster", "")
                if poster:
                    _set("Poster", poster)
            except Exception as e:
                self.info(f"[LABELS] Could not read LABELS file: {e}", "yellow")

        mcf_idx = _col_idx(header_values, "Max Container Fill")
        iw_idx  = _col_idx(header_values, "Item Weight") or _col_idx(header_values, "Accurate Weight")
        if mcf_idx is not None and bag_df is not None and not bag_df.empty:
            items_col = next((c for c in bag_df.columns if str(c).strip().lower() == "items"), None)
            if items_col is not None:
                max_items = bag_df[items_col].map(_to_numeric).dropna().max()
                if max_items is not None:
                    val_row[mcf_idx] = int(max_items)

        for i, hv in enumerate(header_values):
            if str(hv or "") in NUMERIC_COLS:
                val_row[i] = _to_numeric(val_row[i])
            elif str(hv or "") == "Collection Date":
                val_row[i] = _normalise_date(val_row[i])

        if any(v is not None for v in val_row):
            out_ws.append(["Data files"] + val_row)
            row_num = out_ws.max_row
            out_ws.cell(row=row_num, column=1).font = Font(bold=True)

        # Validation row
        validation_row   = [None] * len(header_values)
        validation_fills = {}  # header index → PatternFill

        for col_name in ("Poster", "Client", "Container"):
            idx = _col_idx(header_values, col_name)
            if idx is not None:
                brief_val = str(selected_row[idx] or "").strip()
                data_val  = str(val_row[idx] or "").strip()
                if brief_val and _words_match(brief_val, data_val):
                    validation_row[idx]   = "Correct"
                    validation_fills[idx] = GREEN_FILL
                else:
                    validation_row[idx]   = "Wrong"
                    validation_fills[idx] = RED_FILL

        job_idx = _col_idx(header_values, "Job Name")
        if job_idx is not None:
            brief_val = str(selected_row[job_idx] or "").strip().lower()
            data_val  = str(val_row[job_idx] or "").strip().lower()
            if brief_val and data_val:
                ratio = SequenceMatcher(None, brief_val, data_val).ratio()
                if ratio >= 0.8:
                    validation_row[job_idx]   = "Correct"
                    validation_fills[job_idx] = GREEN_FILL
                else:
                    validation_row[job_idx]   = "Wrong"
                    validation_fills[job_idx] = RED_FILL

        mcf_val_idx = _col_idx(header_values, "Max Container Fill")
        if mcf_val_idx is not None:
            brief_num = _to_numeric(selected_row[mcf_val_idx])
            data_num  = _to_numeric(val_row[mcf_val_idx])
            if brief_num is not None and data_num is not None:
                if data_num <= brief_num:
                    validation_row[mcf_val_idx]   = "Correct"
                    validation_fills[mcf_val_idx] = GREEN_FILL
                else:
                    validation_row[mcf_val_idx]   = "Wrong"
                    validation_fills[mcf_val_idx] = RED_FILL

        if iw_idx is not None:
            brief_num = _to_numeric(selected_row[iw_idx])
            data_num  = _to_numeric(val_row[iw_idx])
            if brief_num is not None and data_num is not None:
                if data_num == brief_num:
                    validation_row[iw_idx]   = "Correct"
                    validation_fills[iw_idx] = GREEN_FILL
                else:
                    validation_row[iw_idx]   = "Wrong"
                    validation_fills[iw_idx] = RED_FILL

        mcw_val_idx = _col_idx(header_values, "Maximum Container Weight")
        if mcw_val_idx is not None:
            brief_num = _to_numeric(selected_row[mcw_val_idx])
            data_num  = _to_numeric(val_row[mcw_val_idx])
            if brief_num is not None and data_num is not None:
                if data_num <= brief_num:
                    validation_row[mcw_val_idx]   = "Correct"
                    validation_fills[mcw_val_idx] = GREEN_FILL
                else:
                    validation_row[mcw_val_idx]   = "Wrong"
                    validation_fills[mcw_val_idx] = RED_FILL

        sr_idx = _col_idx(header_values, "Split Release")
        if sr_idx is not None:
            validation_row[sr_idx]   = "Manual"
            validation_fills[sr_idx] = ORANGE_FILL

        items_idx = _col_idx(header_values, "No. Items")
        if items_idx is not None:
            brief_num = _to_numeric(selected_row[items_idx])
            data_num  = _to_numeric(val_row[items_idx])
            if brief_num is not None and data_num is not None:
                diff = data_num - brief_num
                if 0 <= diff <= 4:
                    validation_row[items_idx]   = "Correct"
                    validation_fills[items_idx] = GREEN_FILL
                else:
                    validation_row[items_idx]   = "Different"
                    validation_fills[items_idx] = ORANGE_FILL

        date_idx = _col_idx(header_values, "Collection Date")
        if date_idx is not None:
            brief_date = str(selected_row[date_idx] or "").strip()
            data_date  = str(val_row[date_idx] or "").strip()
            if brief_date and data_date:
                if brief_date == data_date:
                    validation_row[date_idx]   = "Correct"
                    validation_fills[date_idx] = GREEN_FILL
                else:
                    validation_row[date_idx]   = "Wrong"
                    validation_fills[date_idx] = RED_FILL

        # Service group — scan all four cells together
        svc_indices = [i for i in (_col_idx(header_values, c) for c in GROUP_COL_NAMES)
                       if i is not None]
        group_text = None
        group_fill = None
        if svc_indices:
            brief_text = " ".join(str(selected_row[i] or "") for i in svc_indices)
            data_text  = " ".join(str(val_row[i] or "") for i in svc_indices)
            failed_sections = _validate_service_group(brief_text, data_text)
            if not failed_sections:
                group_text = "Correct"
                group_fill = GREEN_FILL
            else:
                group_text = "Wrong: " + ", ".join(failed_sections)
                group_fill = RED_FILL
            validation_row[svc_indices[0]] = group_text
            for i in svc_indices[1:]:
                validation_row[i] = None
            for i in svc_indices:
                validation_fills[i] = group_fill

        if any(v is not None for v in validation_row):
            out_ws.append(["Validation"] + validation_row)
            row_num = out_ws.max_row
            out_ws.cell(row=row_num, column=1).font = Font(bold=True)
            for col_idx, fill in validation_fills.items():
                cell = out_ws.cell(row=row_num, column=col_idx + 2)
                cell.fill = fill
                cell.font = Font(bold=True)

            if len(svc_indices) > 1:
                min_col = min(svc_indices) + 2
                max_col = max(svc_indices) + 2
                out_ws.merge_cells(
                    start_row=row_num, start_column=min_col,
                    end_row=row_num,   end_column=max_col,
                )
                merged = out_ws.cell(row=row_num, column=min_col)
                merged.value     = group_text
                merged.fill      = group_fill
                merged.font      = Font(bold=True)
                merged.alignment = Alignment(horizontal="center")

        try:
            out_wb.save(out_path)
        except PermissionError:
            raise PermissionError(
                f"Could not save '{os.path.basename(out_path)}'.\n\nPlease close the file in Excel and try again."
            )

    # ------------------------------------------------------------------
    # UI thread: log success and open the output file
    # ------------------------------------------------------------------
    def _on_written(self, out_path: str, job_name: str):
        self.info(f"Job Validation file created for: {job_name}", "green")
        os.startfile(out_path)
